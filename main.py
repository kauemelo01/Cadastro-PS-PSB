"""
Pastoral Social — Consulta de Cadastro
Mobile-first Streamlit app that reads from a GitHub-hosted xlsx file.

Setup:
  1. Upload cadastro.xlsx to your GitHub repo.
  2. Set GITHUB_RAW_URL below (or use st.secrets["GITHUB_RAW_URL"]).
  3. Deploy to Streamlit Community Cloud pointing at this repo.
"""

import streamlit as st
import pandas as pd
import requests
import json
from io import BytesIO
from datetime import datetime, timedelta

# ──────────────────────────────────────────────────────────────
#  CONFIG — Replace with your own GitHub raw URL
#  Format: https://raw.githubusercontent.com/{USER}/{REPO}/{BRANCH}/cadastro.xlsx
# ──────────────────────────────────────────────────────────────
try:
    GITHUB_RAW_URL = st.secrets["GITHUB_RAW_URL"]
except Exception:
    GITHUB_RAW_URL = (
        "https://github.com/kauemelo01/Cadastro-PS-PSB/raw/refs/heads/main/cadastro.xlsx"
    )

# ──────────────────────────────────────────────────────────────
#  PAGE CONFIG
# ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Pastoral Social",
    page_icon="📋",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# ──────────────────────────────────────────────────────────────
#  MOBILE-FIRST CSS
# ──────────────────────────────────────────────────────────────
st.markdown(
    """
<style>
  /* ── Layout ── */
  .main .block-container {
    padding: 0.75rem 0.75rem 2rem;
    max-width: 500px;
    margin: auto;
  }

  /* ── Header ── */
  .ps-header {
    background: linear-gradient(135deg, #1b5e20 0%, #388e3c 100%);
    border-radius: 14px;
    padding: 1rem 1.25rem 0.75rem;
    color: #fff;
    margin-bottom: 1rem;
  }
  .ps-header h1 { margin: 0; font-size: 1.35rem; color: #fff; }
  .ps-header p  { margin: 0.15rem 0 0; font-size: 0.82rem; opacity: 0.85; }

  /* ── Search box — light mode ── */
  .stTextInput > div > div > input {
    font-size: 16px !important;
    border-radius: 12px !important;
    border: 2px solid #c8e6c9 !important;
    padding: 0.65rem 1rem !important;
    background: #f9fbe7 !important;
    color: #212121 !important;
  }
  .stTextInput > div > div > input:focus {
    border-color: #388e3c !important;
    box-shadow: 0 0 0 3px rgba(56,142,60,.15) !important;
  }
  /* ── Search box — dark mode ── */
  @media (prefers-color-scheme: dark) {
    .stTextInput > div > div > input {
      background: #1a2e1a !important;
      color: #e8f5e9 !important;
      border-color: #4caf50 !important;
    }
    .stTextInput > div > div > input::placeholder {
      color: #81c784 !important;
      opacity: 0.7;
    }
    .card { background: #1e2b1e !important; }
    .info-value { color: #e0e0e0 !important; }
    .card-name  { color: #a5d6a7 !important; }
    .card-num   { color: #9e9e9e !important; }
    .sec-title  { color: #9e9e9e !important; }
    .info-grid  { border-top-color: #2e3e2e !important; }
    .info-row   { border-bottom-color: #2a3a2a !important; }
    .info-alert-row { background: #3e2e10 !important; }
    .info-alerta-row { background: #3e1010 !important; }
    .m-no { background: #2a352a !important; }
  }

  /* ── Result card ── */
  .card {
    background: #fff;
    border-radius: 14px;
    padding: 1rem 1.1rem;
    margin-bottom: 0.85rem;
    box-shadow: 0 2px 10px rgba(0,0,0,.08);
    border-left: 5px solid #388e3c;
  }
  .card-name {
    font-size: 1.05rem;
    font-weight: 700;
    color: #1b5e20;
    margin-bottom: 5px;
  }
  .card-num {
    font-size: 0.78rem;
    color: #757575;
    margin-bottom: 8px;
  }

  /* ── Badges ── */
  .badge {
    display: inline-block;
    padding: 3px 11px;
    border-radius: 20px;
    font-size: 0.77rem;
    font-weight: 700;
    margin: 2px 3px 2px 0;
  }
  .b-tipo    { background: #e3f2fd; color: #1565c0; }
  .b-ativo   { background: #e8f5e9; color: #2e7d32; }
  .b-inativo { background: #ffebee; color: #b71c1c; }
  .b-cid     { background: #fce4ec; color: #880e4f; }

  /* ── Section titles inside card ── */
  .sec-title {
    font-size: 0.78rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.04em;
    color: #616161;
    margin: 10px 0 5px;
  }

  /* ── Alert box ── */
  .alert-box {
    background: #fff8e1;
    border: 1px solid #ffe082;
    border-radius: 8px;
    padding: 7px 10px;
    font-size: 0.88rem;
    margin-top: 8px;
    color: #5d4037;
  }

  /* ── Info grid (labeled rows) ── */
  .info-grid {
    margin-top: 10px;
    border-top: 1px solid #f0f0f0;
    padding-top: 8px;
  }
  .info-row {
    display: flex;
    align-items: flex-start;
    padding: 5px 0;
    border-bottom: 1px solid #f9f9f9;
    gap: 8px;
  }
  .info-label {
    font-size: 0.75rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.03em;
    color: #9e9e9e;
    min-width: 80px;
    flex-shrink: 0;
    padding-top: 3px;
  }
  .info-value {
    font-size: 0.9rem;
    color: #212121;
    flex: 1;
  }
  .info-empty { color: #bdbdbd; font-style: italic; }
  .info-alert-row { background: #fff8e1; border-radius: 6px; padding: 5px 7px; margin: 2px 0; }
  .info-alert-val { color: #6d4c41; font-weight: 600; }
  .info-alerta-row { background: #ffebee; border-radius: 6px; padding: 5px 7px; margin: 2px 0; }
  .info-alerta-val { color: #b71c1c; font-weight: 600; font-size: 0.97rem; }

  /* ── Monthly delivery grid ── */
  .month-grid {
    display: grid;
    grid-template-columns: repeat(6, 1fr);
    gap: 4px;
    margin-top: 4px;
  }
  .m-ok {
    background: #e8f5e9;
    border-radius: 6px;
    padding: 3px 2px;
    text-align: center;
    font-size: 0.68rem;
    color: #2e7d32;
    font-weight: 700;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: clip;
  }
  .m-no {
    background: #eeeeee;
    border-radius: 6px;
    min-height: 22px;
    padding: 3px 2px;
  }

  /* ── Search highlight ── */
  mark.hl {
    background: #fff176;
    color: #212121;
    border-radius: 3px;
    padding: 0 2px;
    font-weight: 700;
  }

  /* ── Admin footer (discreet) ── */
  .admin-footer { margin-top: 2.5rem; }
  .admin-badge {
    font-size: 0.78rem;
    color: #2e7d32;
    font-weight: 700;
    padding-top: 8px;
  }
  /* Render the admin link as plain faint text, not a button */
  .st-key-admin_login_link button {
    background: transparent !important;
    border: none !important;
    box-shadow: none !important;
    color: #c5c5c5 !important;
    font-size: 0.72rem !important;
    font-weight: 400 !important;
    padding: 2px 0 !important;
    min-height: 0 !important;
  }
  .st-key-admin_login_link button:hover {
    color: #388e3c !important;
    text-decoration: underline;
  }
  .st-key-admin_login_link button p {
    font-size: 0.72rem !important;
    font-weight: 400 !important;
  }
  .st-key-admin_logout button {
    font-size: 0.75rem !important;
    padding: 2px 8px !important;
    min-height: 0 !important;
  }

  /* ── Empty state ── */
  .empty-state {
    text-align: center;
    padding: 3rem 1rem;
    color: #9e9e9e;
  }
  .empty-state .icon { font-size: 2.8rem; }
  .empty-state p { margin: 0.5rem 0 0; font-size: 0.95rem; }

  /* ── Result count ── */
  .result-count {
    font-size: 0.82rem;
    color: #757575;
    margin-bottom: 0.6rem;
    padding-left: 2px;
  }

  /* ── Hide Streamlit chrome ── */
  #MainMenu { visibility: hidden; }
  footer    { visibility: hidden; }
  header    { visibility: hidden; }
</style>
""",
    unsafe_allow_html=True,
)

# ──────────────────────────────────────────────────────────────
#  HELPERS
# ──────────────────────────────────────────────────────────────
_MONTHS_PT = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
               "Jul", "Ago", "Set", "Out", "Nov", "Dez"]


def serial_to_month(serial: str) -> str:
    """Convert an Excel date serial number string to 'Mmm/YY'."""
    try:
        n = int(float(serial))
        dt = datetime(1899, 12, 30) + timedelta(days=n)
        return f"{_MONTHS_PT[dt.month - 1]}/{str(dt.year)[-2:]}"
    except Exception:
        return str(serial)


def is_numeric(s: str) -> bool:
    try:
        float(s)
        return True
    except (ValueError, TypeError):
        return False


# ──────────────────────────────────────────────────────────────
#  DATA LOADING
# ──────────────────────────────────────────────────────────────
def _fetch_gdrive(url: str) -> bytes:
    """
    Download a file from a Google Drive 'uc?export=download' URL.
    Handles the virus-scan confirmation redirect that Drive adds for larger files.
    """
    session = requests.Session()
    resp = session.get(url, timeout=15)
    resp.raise_for_status()

    # Drive returns an HTML confirmation page when the file is large.
    # Detect it by content-type and extract the confirmed download URL.
    if "text/html" in resp.headers.get("Content-Type", ""):
        # Look for the confirmation token in the response
        import re as _re2
        match = _re2.search(r'confirm=([0-9A-Za-z_\-]+)', resp.text)
        token = match.group(1) if match else "t"
        confirmed_url = url + f"&confirm={token}"
        resp = session.get(confirmed_url, timeout=30)
        resp.raise_for_status()

    return resp.content


@st.cache_data(ttl=300, show_spinner=False)
def load_data(url: str):
    """
    Returns (df, meta, col_index_map, error_msg).
    meta          = {col_name: tag}   tag ∈ {"SEARCH", "Tier 1", "Tier 2"}
    col_index_map = {col_name: 0-based xlsx column index}
    """
    try:
        content = _fetch_gdrive(url)
    except Exception as exc:
        return None, None, {}, str(exc)

    try:
        raw = pd.read_excel(BytesIO(content), sheet_name=0, header=None, dtype=str)
    except Exception as exc:
        return None, None, {}, str(exc)

    if raw.shape[0] < 3:
        return None, None, {}, "Arquivo sem dados suficientes."

    tags_row    = raw.iloc[0].tolist()   # Row 0: SEARCH / Tier 1 / Tier 2
    names_row   = raw.iloc[1].tolist()   # Row 1: human-readable column names

    # Build deduplicated column names
    seen: dict[str, int] = {}
    final_names: list[str] = []
    final_tags:  list[str] = []

    for tag, name in zip(tags_row, names_row):
        raw_name = str(name).strip() if pd.notna(name) else ""
        tag_str  = str(tag).strip()  if pd.notna(tag)  else ""

        # Convert date column names (Excel serial numbers OR datetime strings) to "Mmm/YY"
        if is_numeric(raw_name):
            display_name = serial_to_month(raw_name)
        else:
            # pandas reads Excel dates as "YYYY-MM-DD HH:MM:SS" strings when dtype=str
            try:
                dt = pd.to_datetime(raw_name, errors="raise")
                display_name = f"{_MONTHS_PT[dt.month - 1]}/{str(dt.year)[-2:]}"
            except Exception:
                display_name = raw_name

        # Deduplicate
        if display_name in seen:
            seen[display_name] += 1
            display_name = f"{display_name}_{seen[display_name]}"
        else:
            seen[display_name] = 0

        final_names.append(display_name)
        final_tags.append(tag_str)

    df = raw.iloc[2:].copy()
    df.columns = final_names
    df = df.reset_index(drop=True)

    meta = dict(zip(final_names, final_tags))
    # col_index_map: display_name → 0-based xlsx column index
    col_index_map = {name: idx for idx, name in enumerate(final_names)}
    return df, meta, col_index_map, None


# ──────────────────────────────────────────────────────────────
#  LOAD DATA
# ──────────────────────────────────────────────────────────────
with st.spinner("Carregando cadastro…"):
    df, meta, col_index_map, load_error = load_data(GITHUB_RAW_URL)


def cols_by_tag(tag: str) -> list[str]:
    if not meta:
        return []
    return [c for c, t in meta.items() if t == tag]


SEARCH_COLS = cols_by_tag("SEARCH")
TIER1_COLS  = cols_by_tag("Tier 1")
TIER2_COLS  = cols_by_tag("Tier 2")

# Monthly columns are Tier 1 columns whose names look like "Mmm/YY"
MONTH_COLS  = [c for c in TIER1_COLS if "/" in c and len(c) == 6]
INFO_COLS   = [c for c in TIER1_COLS if c not in MONTH_COLS]


# ──────────────────────────────────────────────────────────────
#  SESSION STATE — in-memory edits overlay
# ──────────────────────────────────────────────────────────────
if "edits" not in st.session_state:
    st.session_state.edits = {}  # {row_idx: {col: value}}

if "is_admin" not in st.session_state:
    st.session_state.is_admin = False


# Admin password — read from secrets only. If unset, admin access is disabled.
try:
    ADMIN_PASSWORD = st.secrets["ADMIN_PASSWORD"]
except Exception:
    ADMIN_PASSWORD = None


def apply_edits(row: pd.Series) -> pd.Series:
    """Return a copy of row with any in-session edits applied."""
    overlay = st.session_state.edits.get(row.name, {})
    if overlay:
        row = row.copy()
        for col, val in overlay.items():
            row[col] = val
    return row


def current_month_col() -> str | None:
    now = datetime.now()
    col = f"{_MONTHS_PT[now.month - 1]}/{str(now.year)[-2:]}"
    return col if col in MONTH_COLS else None


# ──────────────────────────────────────────────────────────────
#  GOOGLE DRIVE WRITE-BACK
# ──────────────────────────────────────────────────────────────
def _gdrive_write_enabled() -> bool:
    """True when both Drive secrets are configured."""
    return (
        "GDRIVE_FILE_ID" in st.secrets
        and "GDRIVE_SA_CREDENTIALS" in st.secrets
    )


def _get_drive_service():
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    creds_info = json.loads(st.secrets["GDRIVE_SA_CREDENTIALS"])
    creds = Credentials.from_service_account_info(
        creds_info,
        scopes=["https://www.googleapis.com/auth/drive"],
    )
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def save_cells_to_drive(row_idx: int, changes: dict[str, str]) -> str | None:
    """
    Download the xlsx from Drive, update one or more cells in a single row,
    then re-upload. One network roundtrip regardless of how many fields changed.
    Returns None on success, or an error string on failure.
    df row_idx is 0-based; xlsx has 2 header rows → xlsx row = row_idx + 3 (1-based).
    """
    if not changes:
        return None
    try:
        from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
        import openpyxl

        file_id = st.secrets["GDRIVE_FILE_ID"]
        service = _get_drive_service()

        # ── Download current file
        req = service.files().get_media(fileId=file_id)
        dl_buf = BytesIO()
        downloader = MediaIoBaseDownload(dl_buf, req)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        dl_buf.seek(0)

        # ── Patch every changed cell
        wb = openpyxl.load_workbook(dl_buf)
        ws = wb.worksheets[0]          # main data always lives on the first sheet
        xlsx_row = row_idx + 3                       # 2 header rows + 1-based index
        for col_name, value in changes.items():
            if col_name not in col_index_map:
                continue
            xlsx_col = col_index_map[col_name] + 1   # 0-based → 1-based
            ws.cell(row=xlsx_row, column=xlsx_col, value=value)

        # ── Upload patched file
        up_buf = BytesIO()
        wb.save(up_buf)
        up_buf.seek(0)
        media = MediaIoBaseUpload(
            up_buf,
            mimetype=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
            resumable=False,
        )
        service.files().update(fileId=file_id, media_body=media).execute()

        # ── Clear cache so next render fetches the fresh file
        load_data.clear()
        return None

    except Exception as exc:
        return str(exc)


def save_cell_to_drive(row_idx: int, col_name: str, value: str) -> str | None:
    """Single-cell convenience wrapper around save_cells_to_drive."""
    return save_cells_to_drive(row_idx, {col_name: value})


# ──────────────────────────────────────────────────────────────
#  CESTAS EXTRAS (Sheet2)
# ──────────────────────────────────────────────────────────────
EXTRAS_SHEET   = "Sheet2"
EXTRAS_HEADERS = ["NÚMERO", "NOME", "CPF", "TELEFONE", "MOTIVO"]
EXTRAS_DONE_COL    = 6          # column F
EXTRAS_DONE_HEADER = "ENTREGUE"


@st.cache_data(ttl=60, show_spinner=False)
def load_extras(url: str) -> pd.DataFrame:
    """Read the Cestas Extras sheet. Returns an empty frame if absent/unreadable."""
    try:
        content = _fetch_gdrive(url)
        extras = pd.read_excel(
            BytesIO(content), sheet_name=EXTRAS_SHEET, dtype=str
        ).fillna("")

        # Normalise the delivered column (F). It may be missing entirely, or
        # present without a header, in which case pandas names it "Unnamed: 5".
        if EXTRAS_DONE_HEADER not in extras.columns:
            if len(extras.columns) >= EXTRAS_DONE_COL:
                extras = extras.rename(
                    columns={extras.columns[EXTRAS_DONE_COL - 1]: EXTRAS_DONE_HEADER}
                )
            else:
                extras[EXTRAS_DONE_HEADER] = ""

        # Drop rows with no NÚMERO and no NOME (blank spacer rows)
        if "NÚMERO" in extras.columns and "NOME" in extras.columns:
            keep = (extras["NÚMERO"].str.strip() != "") | (extras["NOME"].str.strip() != "")
            extras = extras[keep]

        return extras.reset_index(drop=True)
    except Exception:
        return pd.DataFrame(columns=EXTRAS_HEADERS + [EXTRAS_DONE_HEADER])


def next_extra_numero(extras: pd.DataFrame) -> int:
    """Next sequential number for a new Cesta Extra (max existing + 1)."""
    if extras.empty or "NÚMERO" not in extras.columns:
        return 1
    nums = pd.to_numeric(extras["NÚMERO"], errors="coerce").dropna()
    return int(nums.max()) + 1 if len(nums) else 1


def save_extra_to_drive(record: dict[str, str]) -> str | None:
    """
    Append one Cesta Extra row to Sheet2 of the xlsx on Drive.
    Creates the sheet with headers if it does not exist.
    Returns None on success, or an error string on failure.
    """
    try:
        from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
        import openpyxl

        file_id = st.secrets["GDRIVE_FILE_ID"]
        service = _get_drive_service()

        # ── Download current file
        req = service.files().get_media(fileId=file_id)
        dl_buf = BytesIO()
        downloader = MediaIoBaseDownload(dl_buf, req)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        dl_buf.seek(0)

        wb = openpyxl.load_workbook(dl_buf)

        # ── Get or create the extras sheet
        if EXTRAS_SHEET in wb.sheetnames:
            ws = wb[EXTRAS_SHEET]
        else:
            ws = wb.create_sheet(EXTRAS_SHEET)
            for col, header in enumerate(EXTRAS_HEADERS, start=1):
                ws.cell(row=1, column=col, value=header)

        # ── Find the first truly empty row (max_row can overcount blank rows)
        target_row = 2
        for r in range(2, ws.max_row + 2):
            if all(ws.cell(row=r, column=c).value in (None, "")
                   for c in range(1, len(EXTRAS_HEADERS) + 1)):
                target_row = r
                break

        for col, header in enumerate(EXTRAS_HEADERS, start=1):
            ws.cell(row=target_row, column=col, value=record.get(header, ""))

        # ── Upload patched file
        up_buf = BytesIO()
        wb.save(up_buf)
        up_buf.seek(0)
        media = MediaIoBaseUpload(
            up_buf,
            mimetype=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
            resumable=False,
        )
        service.files().update(fileId=file_id, media_body=media).execute()

        load_data.clear()
        load_extras.clear()
        return None

    except Exception as exc:
        return str(exc)


def save_extra_delivered(numero: str) -> str | None:
    """
    Mark a Cesta Extra as delivered by writing "OK" to column F of Sheet2.
    The row is located by matching NÚMERO in column A rather than by position,
    so blank or reordered rows cannot cause the wrong record to be flagged.
    Returns None on success, or an error string on failure.
    """
    try:
        from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
        import openpyxl

        file_id = st.secrets["GDRIVE_FILE_ID"]
        service = _get_drive_service()

        req = service.files().get_media(fileId=file_id)
        dl_buf = BytesIO()
        downloader = MediaIoBaseDownload(dl_buf, req)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        dl_buf.seek(0)

        wb = openpyxl.load_workbook(dl_buf)
        if EXTRAS_SHEET not in wb.sheetnames:
            return f"Aba {EXTRAS_SHEET} não encontrada."
        ws = wb[EXTRAS_SHEET]

        # Ensure column F has a header
        if not ws.cell(row=1, column=EXTRAS_DONE_COL).value:
            ws.cell(row=1, column=EXTRAS_DONE_COL, value=EXTRAS_DONE_HEADER)

        target = None
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val is not None and str(val).strip() == str(numero).strip():
                target = r
                break
        if target is None:
            return f"Registro Nº {numero} não encontrado na aba {EXTRAS_SHEET}."

        ws.cell(row=target, column=EXTRAS_DONE_COL, value="OK")

        up_buf = BytesIO()
        wb.save(up_buf)
        up_buf.seek(0)
        media = MediaIoBaseUpload(
            up_buf,
            mimetype=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
            resumable=False,
        )
        service.files().update(fileId=file_id, media_body=media).execute()

        load_data.clear()
        load_extras.clear()
        return None

    except Exception as exc:
        return str(exc)


# ──────────────────────────────────────────────────────────────
#  DIALOGS
# ──────────────────────────────────────────────────────────────
@st.dialog("Confirmar entrega")
def confirm_month_dialog(row_idx: int, month_col: str, nome: str, numero: str) -> None:
    st.markdown(
        f"Confirmar entrega de **{month_col}** para:<br>"
        f"<b>{nome}</b> &nbsp;·&nbsp; Nº {numero}",
        unsafe_allow_html=True,
    )
    st.write("")
    col1, col2 = st.columns(2)
    if col1.button("✅ Confirmar", use_container_width=True, type="primary"):
        # Update in-session overlay immediately
        st.session_state.edits.setdefault(row_idx, {})[month_col] = "OK"
        # Persist to Drive
        if _gdrive_write_enabled():
            with st.spinner("Salvando no Google Drive…"):
                err = save_cell_to_drive(row_idx, month_col, "OK")
            if err:
                st.error(f"Erro ao salvar: {err}")
                st.stop()
        st.rerun()
    if col2.button("❌ Cancelar", use_container_width=True):
        st.rerun()


@st.dialog("Confirmar entrega")
def confirm_extra_delivery_dialog(numero: str, nome: str) -> None:
    st.markdown(
        f"Confirmar entrega da cesta extra:<br>"
        f"<b>{nome}</b> &nbsp;·&nbsp; Nº {numero}",
        unsafe_allow_html=True,
    )
    st.write("")
    col1, col2 = st.columns(2)
    if col1.button("✅ Confirmar", use_container_width=True, type="primary"):
        if not _gdrive_write_enabled():
            st.error("Salvamento indisponível: credenciais não configuradas.")
            st.stop()
        with st.spinner("Salvando no Google Drive…"):
            err = save_extra_delivered(numero)
        if err:
            st.error(f"Erro ao salvar: {err}")
            st.stop()
        st.session_state.extra_saved = f"Cesta extra Nº {numero} marcada como entregue."
        st.rerun()
    if col2.button("❌ Cancelar", use_container_width=True):
        st.rerun()


@st.dialog("Dar Cesta Extra")
def cesta_extra_dialog(suggested_num: int) -> None:
    st.markdown("Registrar uma nova cesta extra.")

    num = st.text_input(
        "NÚMERO *", value=str(suggested_num), max_chars=10, key="ce_num"
    )
    nome = st.text_input("NOME *", key="ce_nome")
    cpf  = st.text_input("CPF", key="ce_cpf")
    tel  = st.text_input("TELEFONE", key="ce_tel")
    motivo = st.text_area("MOTIVO *", height=90, key="ce_motivo")

    st.caption("* campos obrigatórios")

    # ── Validation
    errors: list[str] = []
    if not num.strip():
        errors.append("NÚMERO é obrigatório.")
    elif not num.strip().isdigit():
        errors.append("NÚMERO deve conter apenas dígitos.")
    if not nome.strip():
        errors.append("NOME é obrigatório.")
    if not motivo.strip():
        errors.append("MOTIVO é obrigatório.")

    st.write("")
    col1, col2 = st.columns(2)
    save_clicked   = col1.button("💾 Salvar",  use_container_width=True, type="primary")
    cancel_clicked = col2.button("❌ Cancelar", use_container_width=True)

    if cancel_clicked:
        st.rerun()

    if save_clicked:
        if errors:
            for e in errors:
                st.error(e)
            st.stop()

        record = {
            "NÚMERO":   num.strip(),
            "NOME":     nome.strip(),
            "CPF":      cpf.strip(),
            "TELEFONE": tel.strip(),
            "MOTIVO":   motivo.strip(),
        }

        if not _gdrive_write_enabled():
            st.error(
                "Salvamento indisponível: credenciais do Google Drive "
                "não configuradas."
            )
            st.stop()

        with st.spinner("Salvando no Google Drive…"):
            err = save_extra_to_drive(record)
        if err:
            st.error(f"Erro ao salvar: {err}")
            st.stop()

        st.session_state.extra_saved = f"Cesta extra Nº {record['NÚMERO']} registrada."
        st.rerun()


@st.dialog("Acesso Administrativo")
def admin_login_dialog() -> None:
    if not ADMIN_PASSWORD:
        st.error(
            "Acesso administrativo indisponível: "
            "`ADMIN_PASSWORD` não está configurada nos Secrets."
        )
        if st.button("Fechar", use_container_width=True):
            st.rerun()
        return

    st.markdown("Informe a senha para habilitar a edição de dados.")
    pwd = st.text_input("Senha", type="password", key="admin_pwd_input")
    st.write("")
    col1, col2 = st.columns(2)
    if col1.button("🔓 Entrar", use_container_width=True, type="primary"):
        if pwd and pwd == ADMIN_PASSWORD:
            st.session_state.is_admin = True
            st.rerun()
        else:
            st.error("Senha incorreta.")
    if col2.button("❌ Cancelar", use_container_width=True):
        st.rerun()


@st.dialog("Editar Dados")
def edit_dados_dialog(row_idx: int, nome: str, current: dict[str, str]) -> None:
    """Edit all Tier 1 editable fields for one record."""
    st.markdown(f"Editando dados de **{nome}**")

    # ── Tipo: dropdown of values already present in the data
    tipo_opts = sorted({v for v in df["TIPO"].fillna("").astype(str).str.strip() if v}) \
        if "TIPO" in df.columns else []
    cur_tipo = current.get("TIPO", "")
    if cur_tipo and cur_tipo not in tipo_opts:
        tipo_opts.insert(0, cur_tipo)
    if "" not in tipo_opts:
        tipo_opts.insert(0, "")
    new_tipo = st.selectbox(
        "Tipo", tipo_opts,
        index=tipo_opts.index(cur_tipo) if cur_tipo in tipo_opts else 0,
    )

    # ── Status: fixed binary
    status_opts = ["", "Ativo", "Inativo"]
    cur_status = current.get("STATUS", "")
    if cur_status and cur_status not in status_opts:
        status_opts.append(cur_status)
    new_status = st.selectbox(
        "Status", status_opts,
        index=status_opts.index(cur_status) if cur_status in status_opts else 0,
    )

    new_cid = st.text_input("CID 2026", value=current.get("CID 2026", ""), max_chars=40)

    st.markdown("**Reservas**")
    new_res1  = st.text_input("Reserva 1",  value=current.get("RESERVA 1", ""))
    new_cpf1  = st.text_input("CPF Res. 1", value=current.get("CPF RESERVA 1", ""))
    new_res2  = st.text_input("Reserva 2",  value=current.get("RESERVA 2", ""))
    new_cpf2  = st.text_input("CPF Res. 2", value=current.get("CPF RESERVA 2", ""))

    new_alerta = st.text_area(
        "⚠️ Alerta para a Mesa",
        value=current.get("ALERTA PARA A MESA", ""),
        height=80,
    )

    proposed = {
        "TIPO":               new_tipo.strip(),
        "STATUS":             new_status.strip(),
        "CID 2026":           new_cid.strip(),
        "RESERVA 1":          new_res1.strip(),
        "CPF RESERVA 1":      new_cpf1.strip(),
        "RESERVA 2":          new_res2.strip(),
        "CPF RESERVA 2":      new_cpf2.strip(),
        "ALERTA PARA A MESA": new_alerta.strip(),
    }
    # Only persist fields that actually changed
    changed = {k: v for k, v in proposed.items() if v != current.get(k, "")}

    st.write("")
    if changed:
        st.caption(f"{len(changed)} campo(s) alterado(s): " + ", ".join(changed.keys()))
    else:
        st.caption("Nenhuma alteração.")

    col1, col2 = st.columns(2)
    if col1.button("💾 Salvar", use_container_width=True, type="primary",
                   disabled=not changed):
        st.session_state.edits.setdefault(row_idx, {}).update(changed)
        if _gdrive_write_enabled():
            with st.spinner("Salvando no Google Drive…"):
                err = save_cells_to_drive(row_idx, changed)
            if err:
                st.error(f"Erro ao salvar: {err}")
                st.stop()
        st.rerun()
    if col2.button("❌ Cancelar", use_container_width=True):
        st.rerun()


# ──────────────────────────────────────────────────────────────
#  SEARCH FUNCTION
# ──────────────────────────────────────────────────────────────
import unicodedata as _ud

_CPF_COLS = {"CPF", "CPF RESERVA 1", "CPF RESERVA 2"}

# Map each base vowel/ç/ñ (and its accented variants) to a regex char-class
# so that searching "Antonio" also matches "Antônio" and vice-versa.
_ACCENT_CLASS: dict[str, str] = {}
for _base, _variants in [
    ("a", "aàáâãäå"), ("e", "eèéêë"), ("i", "iìíîï"),
    ("o", "oòóôõö"), ("u", "uùúûü"), ("c", "cç"),  ("n", "nñ"),
]:
    _cls = f"[{_variants}]"
    for _ch in _variants:
        _ACCENT_CLASS[_ch] = _cls


def _normalize(s: str) -> str:
    """Lowercase + strip all combining diacritical marks."""
    return "".join(
        c for c in _ud.normalize("NFD", s) if _ud.category(c) != "Mn"
    ).lower()


def _strip_cpf(s: str) -> str:
    """Remove dots and dashes so raw and formatted CPFs compare equal."""
    return s.replace(".", "").replace("-", "")


def search_numero_exact(query: str) -> pd.DataFrame:
    """Return only the row whose NUMERO exactly matches the query."""
    q = query.strip()
    if not q or df is None or "NUMERO" not in df.columns:
        return pd.DataFrame()
    return df[df["NUMERO"].fillna("").str.strip() == q].copy()


def search_df(query: str, cols: list[str] | None = None) -> pd.DataFrame:
    q_norm = _normalize(query.strip())
    if not q_norm or df is None:
        return pd.DataFrame()
    if cols is None:
        cols = SEARCH_COLS
    q_cpf = _strip_cpf(q_norm)
    mask = pd.Series([False] * len(df), index=df.index)
    for col in cols:
        if col not in df.columns:
            continue
        if col in _CPF_COLS:
            mask |= (
                df[col].fillna("").apply(lambda v: _strip_cpf(_normalize(v)))
                .str.contains(q_cpf, regex=False)
            )
        else:
            mask |= (
                df[col].fillna("").apply(_normalize)
                .str.contains(q_norm, regex=False)
            )
    return df[mask].copy()


# ──────────────────────────────────────────────────────────────
#  RENDER RECORD
# ──────────────────────────────────────────────────────────────
def cell(row: pd.Series, col: str) -> str:
    v = str(row.get(col, "")).strip()
    return "" if v in ("nan", "None", "") else v


import re as _re


def _query_to_pattern(query: str) -> str:
    """Build a regex pattern from query that matches accented variants."""
    parts = []
    for ch in query.lower():
        if ch in _ACCENT_CLASS:
            parts.append(_ACCENT_CLASS[ch])
        else:
            parts.append(_re.escape(ch))
    return "".join(parts)


def highlight(text: str, query: str) -> str:
    """Highlight query in text, accent-insensitive."""
    if not query or not text:
        return text
    try:
        pattern = _query_to_pattern(query)
        return _re.sub(f"({pattern})", r'<mark class="hl">\1</mark>', text, flags=_re.IGNORECASE)
    except _re.error:
        return text


def highlight_cpf(text: str, query: str) -> str:
    """Highlight a CPF field, matching regardless of dots/dashes formatting."""
    if not query or not text:
        return text
    result = highlight(text, query)
    if result != text:
        return result
    if _strip_cpf(_normalize(query)) in _strip_cpf(_normalize(text)):
        return f'<mark class="hl">{text}</mark>'
    return text


def sort_results(results: pd.DataFrame, query: str) -> pd.DataFrame:
    """Push rows where NOME or CPF match the query to the top."""
    if results.empty or not query:
        return results
    q_norm = _normalize(query.strip())
    q_cpf  = _strip_cpf(q_norm)

    def priority(row: pd.Series) -> int:
        if q_norm in _normalize(str(row.get("NOME", ""))):
            return 0
        if q_cpf in _strip_cpf(_normalize(str(row.get("CPF", "")))):
            return 0
        return 1

    results = results.copy()
    results["_pri"] = results.apply(priority, axis=1)
    return results.sort_values("_pri", kind="stable").drop(columns=["_pri"])


def render_record(row: pd.Series, query: str = "", numero_query: str = "") -> None:
    row = apply_edits(row)
    row_idx = row.name

    q     = query.strip().lower()
    q_num = numero_query.strip()

    nome   = cell(row, "NOME")
    numero = cell(row, "NUMERO")
    cpf    = cell(row, "CPF")
    tipo   = cell(row, "TIPO")
    status = cell(row, "STATUS")
    cid    = cell(row, "CID 2026")
    alerta = cell(row, "ALERTA PARA A MESA")
    reserva1     = cell(row, "RESERVA 1")
    cpf_reserva1 = cell(row, "CPF RESERVA 1")
    reserva2     = cell(row, "RESERVA 2")
    cpf_reserva2 = cell(row, "CPF RESERVA 2")

    status_cls = "b-ativo" if status.lower() == "ativo" else "b-inativo"

    # ── Header: name + number + CPF
    # NUMERO highlights via numero_query (NUMERO-bar search) or general query
    numero_hl = highlight(numero, q_num) if q_num else highlight(numero, q)
    html = f"""
    <div class="card">
      <div class="card-name">{highlight(nome, q) or "—"}</div>
      <div class="card-num">Nº {numero_hl} &nbsp;·&nbsp; CPF {highlight_cpf(cpf, q) if cpf else "—"}</div>
    """

    # ── Info rows: always show all Tier 1 fields explicitly
    html += '<div class="info-grid">'

    html += f"""
      <div class="info-row">
        <span class="info-label">Tipo</span>
        <span class="info-value"><span class="badge b-tipo">{tipo or "—"}</span></span>
      </div>
      <div class="info-row">
        <span class="info-label">Status</span>
        <span class="info-value"><span class="badge {status_cls}">{status or "—"}</span></span>
      </div>
      <div class="info-row">
        <span class="info-label">CID 2026</span>
        <span class="info-value">{highlight(cid, q) if cid else '<span class="info-empty">—</span>'}</span>
      </div>
    """

    def reserva_row(label: str, val: str, is_cpf: bool = False) -> str:
        if val:
            v_hl = highlight_cpf(val, q) if is_cpf else highlight(val, q)
            matched = (
                (_strip_cpf(q) in _strip_cpf(val.lower())) if is_cpf
                else (q in val.lower())
            )
        else:
            v_hl = '<span class="info-empty">—</span>'
            matched = False
        row_cls = ' info-alert-row' if matched else ''
        return f'''
      <div class="info-row{row_cls}">
        <span class="info-label">{label}</span>
        <span class="info-value">{v_hl}</span>
      </div>'''

    html += reserva_row("Reserva 1",   reserva1)
    html += reserva_row("CPF Res. 1",  cpf_reserva1, is_cpf=True)
    html += reserva_row("Reserva 2",   reserva2)
    html += reserva_row("CPF Res. 2",  cpf_reserva2, is_cpf=True)

    # ── Alert row — always visible
    if alerta and alerta not in ("0",):
        html += f"""
      <div class="info-row info-alerta-row">
        <span class="info-label">⚠️ Alerta</span>
        <span class="info-value info-alerta-val">{highlight(alerta, q)}</span>
      </div>
        """
    else:
        html += """
      <div class="info-row">
        <span class="info-label">Alerta</span>
        <span class="info-value info-empty">Nenhum</span>
      </div>
        """

    # ── Any other non-monthly Tier 1 columns not yet handled above
    already_shown = {"TIPO", "CID 2026", "STATUS", "ALERTA PARA A MESA",
                     "RESERVA 1", "CPF RESERVA 1", "RESERVA 2", "CPF RESERVA 2"}
    for col in INFO_COLS:
        if col not in already_shown:
            val = cell(row, col)
            html += f"""
      <div class="info-row">
        <span class="info-label">{col}</span>
        <span class="info-value">{highlight(val, q) if val else '<span class="info-empty">—</span>'}</span>
      </div>
            """

    html += "</div>"  # close info-grid

    # ── Monthly deliveries — compact grid: OK = green cell, blank = empty dot
    html += '<div class="sec-title">📦 Histórico de Entregas</div>'
    html += '<div class="month-grid">'
    for mc in MONTH_COLS:
        val = cell(row, mc)
        if val.lower() == "ok":
            html += f'<div class="m-ok" title="{mc}">✓ {mc}</div>'
        else:
            html += f'<div class="m-no" title="{mc}"></div>'
    html += "</div>"

    html += "</div>"   # close card

    st.markdown(html, unsafe_allow_html=True)

    # ── Action buttons
    cur_month = current_month_col()
    month_ok  = cur_month and cell(row, cur_month).lower() == "ok"

    btn_cols = st.columns(2)
    with btn_cols[0]:
        is_admin = st.session_state.get("is_admin", False)
        if st.button(
            "✏️ Editar Dados" if is_admin else "🔒 Editar Dados",
            key=f"cid_{row_idx}",
            use_container_width=True,
            disabled=not is_admin,
            help=None if is_admin else "Requer acesso administrativo",
        ):
            edit_dados_dialog(row_idx, nome, {
                "TIPO":               tipo,
                "STATUS":             status,
                "CID 2026":           cid,
                "RESERVA 1":          reserva1,
                "CPF RESERVA 1":      cpf_reserva1,
                "RESERVA 2":          reserva2,
                "CPF RESERVA 2":      cpf_reserva2,
                "ALERTA PARA A MESA": alerta,
            })
    with btn_cols[1]:
        if cur_month:
            if month_ok:
                st.button(
                    f"✅ {cur_month} OK",
                    key=f"flag_{row_idx}",
                    use_container_width=True,
                    disabled=True,
                )
            else:
                if st.button(
                    f"📦 Marcar {cur_month}",
                    key=f"flag_{row_idx}",
                    use_container_width=True,
                ):
                    confirm_month_dialog(row_idx, cur_month, nome, numero)


# ──────────────────────────────────────────────────────────────
#  SCREEN — LISTA DE CESTAS EXTRAS
# ──────────────────────────────────────────────────────────────
def _render_extras_list() -> None:
    """Body of the Cestas Extras screen. Re-runs on its own every 60s."""
    extras = load_extras(GITHUB_RAW_URL)

    st.caption(
        f"Atualizado às {datetime.now().strftime('%H:%M:%S')} · "
        "atualização automática a cada minuto"
    )

    if extras.empty:
        st.markdown(
            """
<div class="empty-state">
  <div class="icon">🧺</div>
  <p>Nenhuma cesta extra registrada ainda.</p>
</div>
""",
            unsafe_allow_html=True,
        )
        return

    total     = len(extras)
    entregues = (
        extras[EXTRAS_DONE_HEADER].str.strip().str.lower() == "ok"
    ).sum()
    st.markdown(
        f'<div class="result-count">{total} cesta(s) extra(s) · '
        f'{entregues} entregue(s) · {total - entregues} pendente(s)</div>',
        unsafe_allow_html=True,
    )

    for i, row in extras.iterrows():
        numero = str(row.get("NÚMERO", "")).strip()
        nome   = str(row.get("NOME", "")).strip()
        cpf    = str(row.get("CPF", "")).strip()
        tel    = str(row.get("TELEFONE", "")).strip()
        motivo = str(row.get("MOTIVO", "")).strip()
        done   = str(row.get(EXTRAS_DONE_HEADER, "")).strip().lower() == "ok"

        badge = (
            '<span class="badge b-ativo">ENTREGUE</span>' if done
            else '<span class="badge b-inativo">PENDENTE</span>'
        )

        html = f"""
        <div class="card">
          <div class="card-name">{nome or "—"}</div>
          <div class="card-num">Nº {numero or "—"}
            &nbsp;·&nbsp; CPF {cpf or "—"}
            &nbsp;·&nbsp; Tel {tel or "—"}</div>
          <div class="info-grid">
            <div class="info-row">
              <span class="info-label">Situação</span>
              <span class="info-value">{badge}</span>
            </div>
            <div class="info-row">
              <span class="info-label">Motivo</span>
              <span class="info-value">{motivo or '<span class="info-empty">—</span>'}</span>
            </div>
          </div>
        </div>
        """
        st.markdown(html, unsafe_allow_html=True)

        if done:
            st.button(
                "✅ Entregue",
                key=f"extra_done_{i}_{numero}",
                use_container_width=True,
                disabled=True,
            )
        else:
            if st.button(
                "📦 Marcar como entregue",
                key=f"extra_done_{i}_{numero}",
                use_container_width=True,
            ):
                confirm_extra_delivery_dialog(numero, nome)


# Auto-refresh every 60s where supported, otherwise render normally.
try:
    render_extras_list = st.fragment(run_every="60s")(_render_extras_list)
except Exception:
    render_extras_list = _render_extras_list


# ──────────────────────────────────────────────────────────────
#  UI — HEADER
# ──────────────────────────────────────────────────────────────
st.markdown(
    """
<div class="ps-header">
  <h1>📋 Pastoral Social</h1>
  <p>Consulta de Cadastro · Jardim Colombo</p>
</div>
""",
    unsafe_allow_html=True,
)

# Show Drive write status — only when secrets are missing
if not _gdrive_write_enabled():
    st.warning(
        "⚠️ Salvamento automático desativado. "
        "Configure `GDRIVE_FILE_ID` e `GDRIVE_SA_CREDENTIALS` nos Secrets para habilitar.",
        icon="🔒",
    )

# ──────────────────────────────────────────────────────────────
#  UI — ERROR STATE
# ──────────────────────────────────────────────────────────────
if load_error or df is None:
    st.error("❌ Não foi possível carregar o arquivo.")
    st.markdown(
        """
**Possíveis causas:**
- A URL do arquivo no GitHub ainda não foi configurada.
- O repositório ou arquivo não existe / é privado.
- Problema de rede.

**Solução:** edite `app.py` e defina `GITHUB_RAW_URL` com a URL raw do seu arquivo,
ou crie o arquivo `secrets.toml` com a chave `GITHUB_RAW_URL`.
"""
    )
    if load_error:
        with st.expander("Detalhes do erro"):
            st.code(load_error)
    st.stop()

# ──────────────────────────────────────────────────────────────
#  SCREEN ROUTING — Cestas Extras (admin only)
# ──────────────────────────────────────────────────────────────
if "view" not in st.session_state:
    st.session_state.view = "main"

# Losing admin must never leave the user stranded on a restricted screen
if st.session_state.view == "extras" and not st.session_state.get("is_admin", False):
    st.session_state.view = "main"

if st.session_state.view == "extras":
    back, _sp = st.columns([1, 2])
    if back.button("← Voltar", key="extras_back", use_container_width=True):
        st.session_state.view = "main"
        st.rerun()

    if st.session_state.get("extra_saved"):
        st.success(f"✅ {st.session_state.pop('extra_saved')}")

    st.markdown('<div class="sec-title">🧺 Lista de Cestas Extras</div>',
                unsafe_allow_html=True)
    render_extras_list()
    st.stop()

# ──────────────────────────────────────────────────────────────
#  UI — SEARCH
# ──────────────────────────────────────────────────────────────
query_num = st.text_input(
    "Buscar por Número",
    placeholder="🔢  Número do registro…",
    label_visibility="collapsed",
    key="q_num",
)
query_gen = st.text_input(
    "Buscar geral",
    placeholder="🔍  Nome, CPF, reserva…",
    label_visibility="collapsed",
    key="q_gen",
)


def show_results(
    num_results: pd.DataFrame,
    gen_results: pd.DataFrame,
    num_query: str,
    gen_query: str,
) -> None:
    # OR logic: union of both result sets; gen_results take highlight priority
    gen_idx = set(gen_results.index) if not gen_results.empty else set()
    num_idx = set(num_results.index) if not num_results.empty else set()
    all_idx = gen_idx | num_idx

    if not all_idx:
        st.markdown(
            """
<div class="empty-state">
  <div class="icon">🔎</div>
  <p>Nenhum registro encontrado.<br>
  Tente outro nome, CPF ou número.</p>
</div>
""",
            unsafe_allow_html=True,
        )
        return

    # Preserve original df order; sort within that by NOME/CPF priority
    combined = df.loc[sorted(all_idx)].copy()
    # Use gen_query for sorting if available, else num_query
    combined = sort_results(combined, gen_query or num_query)

    count = len(combined)
    st.markdown(
        f'<div class="result-count">{count} registro{"s" if count > 1 else ""} encontrado{"s" if count > 1 else ""}</div>',
        unsafe_allow_html=True,
    )
    for idx, row in combined.iterrows():
        if idx in gen_idx:
            # General match: highlight all matching fields
            render_record(row, query=gen_query, numero_query="")
        else:
            # NUMERO-only match: highlight only the NUMERO header
            render_record(row, query="", numero_query=num_query)


# ──────────────────────────────────────────────────────────────
#  UI — RESULTS
# ──────────────────────────────────────────────────────────────
num_results = search_numero_exact(query_num) if query_num else pd.DataFrame()
gen_results = search_df(query_gen)           if query_gen else pd.DataFrame()

if query_num or query_gen:
    show_results(num_results, gen_results, query_num, query_gen)
else:
    st.markdown(
        f"""
<div class="empty-state">
  <div class="icon">🔍</div>
  <p>Digite um número de registro<br>ou busque por nome, CPF ou reserva.</p>
  <p style="margin-top:1.5rem;font-size:0.8rem;color:#bdbdbd">
    {len(df)} registros · {len(SEARCH_COLS)} campos de busca
  </p>
</div>
""",
        unsafe_allow_html=True,
    )

# ──────────────────────────────────────────────────────────────
#  UI — ADMIN FOOTER (discreet)
# ──────────────────────────────────────────────────────────────
st.markdown('<div class="admin-footer">', unsafe_allow_html=True)

if st.session_state.get("is_admin", False):
    # ── Success toast from a previous save
    if st.session_state.get("extra_saved"):
        st.success(f"✅ {st.session_state.pop('extra_saved')}")

    ecol1, ecol2 = st.columns(2)
    if ecol1.button("🧺 Dar Cesta Extra", key="btn_cesta_extra",
                    use_container_width=True, type="primary"):
        extras = load_extras(GITHUB_RAW_URL)
        cesta_extra_dialog(next_extra_numero(extras))
    if ecol2.button("📋 Lista de Cestas Extras", key="btn_lista_extras",
                    use_container_width=True):
        st.session_state.view = "extras"
        st.rerun()

    fcol1, fcol2 = st.columns([3, 2])
    fcol1.markdown(
        '<div class="admin-badge">🔓 Modo administrador ativo</div>',
        unsafe_allow_html=True,
    )
    if fcol2.button("Sair", key="admin_logout", use_container_width=True):
        st.session_state.is_admin = False
        st.rerun()
else:
    if st.button("· acesso administrativo ·", key="admin_login_link"):
        admin_login_dialog()

st.markdown("</div>", unsafe_allow_html=True)
